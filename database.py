# -*- coding: utf-8 -*-
import sqlite3
import json
from contextlib import contextmanager
from pathlib import Path

from config import Config


def get_db_path():
    if Config.DATABASE_PATH:
        return Config.DATABASE_PATH
    return str(Path(Config.BASE_DIR) / "vetguardian.db")

@contextmanager
def get_connection():
    conn = sqlite3.connect(get_db_path())
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()

def init_db():
    print("[DB] init_db() start")
    with get_connection() as conn:
        print("[DB] creating tables if not exist")
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            name TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS pets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name TEXT,
            species TEXT,
            breed TEXT,
            age_group TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (user_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS cases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            pet_id INTEGER,
            symptoms_data TEXT,
            danger_level TEXT,
            result_summary TEXT,
            result_details TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (user_id) REFERENCES users(id),
            FOREIGN KEY (pet_id) REFERENCES pets(id)
        );

        -- Статьи и статистика по породам
        CREATE TABLE IF NOT EXISTS breed_stats (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            breed_name TEXT NOT NULL UNIQUE,
            description TEXT,
            common_issues TEXT,
            typical_diseases TEXT,  -- JSON массив болезней
            disease_frequency TEXT, -- JSON объект с частотами
            trait_frequency TEXT,   -- JSON объект с особенностями
            total_cases INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        -- Статьи и статистика по возрасту
        CREATE TABLE IF NOT EXISTS age_stats (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            age_group TEXT NOT NULL UNIQUE,
            description TEXT,
            care_recommendations TEXT,
            common_problems TEXT,
            complications_frequency TEXT, -- JSON объект
            diseases_by_care TEXT,        -- JSON объект
            total_cases INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        -- Статьи и статистика по поведению
        CREATE TABLE IF NOT EXISTS behavior_stats (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            behavior_type TEXT NOT NULL UNIQUE,
            description TEXT,
            causes TEXT,
            solutions TEXT,
            frequency INTEGER DEFAULT 0,
            total_cases INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        -- Хранение сырых данных анамнеза для статистики
        CREATE TABLE IF NOT EXISTS anamnesis_cases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            breed TEXT,
            age_group TEXT,
            symptoms TEXT,  -- JSON массив симптомов
            behaviors TEXT, -- JSON массив поведенческих проблем
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)

        print("[DB] inserting default reference data (if missing)")
        # Начальная загрузка справочных данных по возрастам (породы и поведение берём из Excel)
        # Возрастные группы
        conn.execute("""
        INSERT OR IGNORE INTO age_stats
            (age_group, description, care_recommendations, common_problems, complications_frequency, diseases_by_care)
        VALUES
            ('Щенки (0-1 год)',
             'Активный рост, формирование скелета, смена зубов, социализация.',
             'Частое кормление, вакцинация по графику, ранняя социализация.',
             'Неправильный прикус, инфекции, травмы из-за гиперактивности.',
             json('{"Энтерит":40,"Чума":15,"Травмы":50}'),
             json('{"Рахит":30,"Ожирение":25,"Неправильный прикус":40}')
            );
        """)
        conn.execute("""
        INSERT OR IGNORE INTO age_stats
            (age_group, description, care_recommendations, common_problems, complications_frequency, diseases_by_care)
        VALUES
            ('Молодые (1-3 года)',
             'Пик физической формы, половое созревание, закрепление поведения.',
             'Регулярные нагрузки, дрессировка, контроль веса.',
             'Травмы, агрессия (у кобелей), аллергии.',
             json('{"Травмы":60,"Отравления":35,"Заворот желудка":10}'),
             json('{"Ожирение":40,"Артрит":15,"Аллергии":45}')
            );
        """)
        conn.execute("""
        INSERT OR IGNORE INTO age_stats
            (age_group, description, care_recommendations, common_problems, complications_frequency, diseases_by_care)
        VALUES
            ('Взрослые (3-7 лет)',
             'Стабильное состояние, зрелость, возможно начало хронических заболеваний.',
             'Профилактические осмотры, контроль зубного камня.',
             'Зубной камень, ожирение, начало артрита.',
             json('{"Артрит":50,"Мочекаменная болезнь":30,"Диабет":20}'),
             json('{"Ожирение":55,"Зубной камень":70,"Гиподинамия":40}')
            );
        """)
        conn.execute("""
        INSERT OR IGNORE INTO age_stats
            (age_group, description, care_recommendations, common_problems, complications_frequency, diseases_by_care)
        VALUES
            ('Пожилые (7+ лет)',
             'Замедление обмена веществ, возрастные изменения органов.',
             'Легкая пища, добавки для суставов, частые осмотры.',
             'Артрит, ухудшение зрения/слуха, недержание.',
             json('{"Артрит":80,"Почечная недостаточность":40,"Катаракта":50}'),
             json('{"Ожирение":45,"Сердечная недостаточность":55,"Онкология":30}')
            );
        """)

        print("[DB] running soft migrations")
        # Мягкие миграции для уже существующих БД.
        # 1) добавляем недостающие колонки, чтобы эндпоинты статистики не падали
        #    на старых схемах (отсутствуют поля complications_frequency и т.п.);
        # 2) для уже существующих строк, где новые поля NULL/пустые, проставляем
        #    базовые справочные значения, как при первоначальной инициализации.
        def ensure_column(table, name, ddl):
            cols = {row["name"] for row in conn.execute(f"PRAGMA table_info({table})")}
            if name not in cols:
                conn.execute(f"ALTER TABLE {table} ADD COLUMN {ddl}")

        # age_stats: новые поля для статистики осложнений и ухода
        ensure_column("age_stats", "complications_frequency", "complications_frequency TEXT")
        ensure_column("age_stats", "diseases_by_care", "diseases_by_care TEXT")
        ensure_column("age_stats", "total_cases", "total_cases INTEGER DEFAULT 0")
        ensure_column("age_stats", "created_at", "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP")
        ensure_column("age_stats", "updated_at", "updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP")

        # behavior_stats: частота, счётчик кейсов и временные метки
        ensure_column("behavior_stats", "frequency", "frequency INTEGER DEFAULT 0")
        ensure_column("behavior_stats", "total_cases", "total_cases INTEGER DEFAULT 0")
        ensure_column("behavior_stats", "created_at", "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP")
        ensure_column("behavior_stats", "updated_at", "updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP")

        # Если таблица age_stats уже существовала до добавления новых полей,
        # проставляем для известных возрастных групп базовые частоты,
        # но только там, где ещё нет своих значений.
        conn.execute(
            """
            UPDATE age_stats
            SET complications_frequency = COALESCE(
                    NULLIF(complications_frequency, ''),
                    json('{"Энтерит":40,"Чума":15,"Травмы":50}')
                ),
                diseases_by_care = COALESCE(
                    NULLIF(diseases_by_care, ''),
                    json('{"Рахит":30,"Ожирение":25,"Неправильный прикус":40}')
                )
            WHERE age_group = 'Щенки (0-1 год)'
            """
        )
        conn.execute(
            """
            UPDATE age_stats
            SET complications_frequency = COALESCE(
                    NULLIF(complications_frequency, ''),
                    json('{"Травмы":60,"Отравления":35,"Заворот желудка":10}')
                ),
                diseases_by_care = COALESCE(
                    NULLIF(diseases_by_care, ''),
                    json('{"Ожирение":40,"Артрит":15,"Аллергии":45}')
                )
            WHERE age_group = 'Молодые (1-3 года)'
            """
        )
        conn.execute(
            """
            UPDATE age_stats
            SET complications_frequency = COALESCE(
                    NULLIF(complications_frequency, ''),
                    json('{"Артрит":50,"Мочекаменная болезнь":30,"Диабет":20}')
                ),
                diseases_by_care = COALESCE(
                    NULLIF(diseases_by_care, ''),
                    json('{"Ожирение":55,"Зубной камень":70,"Гиподинамия":40}')
                )
            WHERE age_group = 'Взрослые (3-7 лет)'
            """
        )
        conn.execute(
            """
            UPDATE age_stats
            SET complications_frequency = COALESCE(
                    NULLIF(complications_frequency, ''),
                    json('{"Артрит":80,"Почечная недостаточность":40,"Катаракта":50}')
                ),
                diseases_by_care = COALESCE(
                    NULLIF(diseases_by_care, ''),
                    json('{"Ожирение":45,"Сердечная недостаточность":55,"Онкология":30}')
                )
            WHERE age_group = 'Пожилые (7+ лет)'
            """
        )

        # После всех базовых INSERT/UPDATE пробуем подтянуть
        # актуальные статьи и статистику из Excel-файла, если он есть.
        print("[DB] trying to load stats from Excel")
        _load_stats_from_excel(conn)
        print("[DB] init_db() done")


def _load_stats_from_excel(conn):
    """
    Загружает статьи и статистику из Excel-файла articles_and_stats.xlsx,
    если он присутствует рядом с проектом. Не является обязательным:
    при отсутствии файла или библиотеки openpyxl функция просто завершится,
    не ломая приложение.
    """
    print("[DB] _load_stats_from_excel() called")
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        print("[DB] openpyxl not available, skip Excel load")
        # Нет openpyxl — тихо выходим, остаются значения, зашитые в init_db()
        return

    # Ищем Excel рядом с приложением (корень проекта VetGuardian)
    xlsx_path = Path(__file__).resolve().parent / "articles_and_stats.xlsx"
    print("[DB] Excel path:", xlsx_path, "exists:", xlsx_path.exists())
    if not xlsx_path.exists():
        print("[DB] Excel file not found, skip")
        return

    try:
        wb = load_workbook(filename=str(xlsx_path), data_only=True)
    except Exception:
        print("[DB] failed to load workbook, skip")
        return

    def sheet_to_rows(ws):
        """Преобразует лист в список словарей по заголовкам первой строки."""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print("[DB] sheet", ws.title, "is empty")
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        result = []
        for row in rows[1:]:
            item = {}
            for idx, val in enumerate(row):
                key = headers[idx] if idx < len(headers) else ""
                if not key:
                    continue
                item[key] = val
            # пропускаем полностью пустые строки
            if any(v not in (None, "", " ") for v in item.values()):
                result.append(item)
        return result

    # --- Лист по породам ---
    print("[DB] workbook sheets:", wb.sheetnames)

    seen_breeds = set()
    if "Breeds" in wb.sheetnames:
        ws = wb["Breeds"]
        rows = sheet_to_rows(ws)
        print("[DB] Breeds rows:", len(rows))
        for row in rows:
            breed_name = (row.get("breed_name") or "").strip()
            if not breed_name:
                continue
            seen_breeds.add(breed_name)
            description = (row.get("description") or "") or None
            common_issues = (row.get("common_issues") or "") or None

            # Типичные заболевания – список через ;
            td_raw = row.get("typical_diseases") or ""
            typical_diseases = [
                part.strip() for part in str(td_raw).split(";") if str(part).strip()
            ]

            # График заболеваний
            dl_raw = row.get("disease_labels") or ""
            dv_raw = row.get("disease_values") or ""
            disease_labels = [
                part.strip() for part in str(dl_raw).split(";") if str(part).strip()
            ]
            disease_values = [
                part.strip() for part in str(dv_raw).split(";") if str(part).strip()
            ]
            disease_freq = {}
            for label, value in zip(disease_labels, disease_values):
                try:
                    num = float(str(value).replace(",", ".") or "0")
                except Exception:
                    num = 0.0
                # Если в Excel указаны доли (0.4), преобразуем в проценты
                if 0 < num <= 1:
                    num *= 100.0
                disease_freq[label] = num

            # График особенностей
            tl_raw = row.get("trait_labels") or ""
            tv_raw = row.get("trait_values") or ""
            trait_labels = [
                part.strip() for part in str(tl_raw).split(";") if str(part).strip()
            ]
            trait_values = [
                part.strip() for part in str(tv_raw).split(";") if str(part).strip()
            ]
            trait_freq = {}
            for label, value in zip(trait_labels, trait_values):
                try:
                    num = float(str(value).replace(",", ".") or "0")
                except Exception:
                    num = 0.0
                if 0 < num <= 1:
                    num *= 100.0
                trait_freq[label] = num

            conn.execute(
                """
                INSERT INTO breed_stats
                    (breed_name, description, common_issues,
                     typical_diseases, disease_frequency, trait_frequency)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(breed_name) DO UPDATE SET
                    description = excluded.description,
                    common_issues = excluded.common_issues,
                    typical_diseases = excluded.typical_diseases,
                    disease_frequency = excluded.disease_frequency,
                    trait_frequency = excluded.trait_frequency,
                    updated_at = CURRENT_TIMESTAMP
                """,
                (
                    breed_name,
                    description,
                    common_issues,
                    json.dumps(typical_diseases, ensure_ascii=False),
                    json.dumps(disease_freq, ensure_ascii=False),
                    json.dumps(trait_freq, ensure_ascii=False),
                ),
            )

    # Если удалось прочитать хоть одну породу из Excel —
    # чистим таблицу breed_stats от записей, которых нет в Excel.
    if seen_breeds:
        placeholders = ",".join("?" for _ in seen_breeds)
        conn.execute(
            f"DELETE FROM breed_stats WHERE breed_name NOT IN ({placeholders})",
            tuple(seen_breeds),
        )

    # --- Лист по возрасту ---
    if "AgeGroups" in wb.sheetnames:
        ws = wb["AgeGroups"]
        rows = sheet_to_rows(ws)
        print("[DB] AgeGroups rows:", len(rows))
        for row in rows:
            age_group = (row.get("age_group") or "").strip()
            if not age_group:
                continue
            description = (row.get("description") or "") or None
            care_recommendations = (row.get("care_recommendations") or "") or None
            common_problems = (row.get("common_problems") or "") or None

            cl_raw = row.get("complications_labels") or ""
            cv_raw = row.get("complications_values") or ""
            comp_labels = [
                part.strip() for part in str(cl_raw).split(";") if str(part).strip()
            ]
            comp_values = [
                part.strip() for part in str(cv_raw).split(";") if str(part).strip()
            ]
            comp_freq = {}
            for label, value in zip(comp_labels, comp_values):
                try:
                    num = float(str(value).replace(",", ".") or "0")
                except Exception:
                    num = 0.0
                if 0 < num <= 1:
                    num *= 100.0
                comp_freq[label] = num

            crl_raw = row.get("care_risk_labels") or ""
            crv_raw = row.get("care_risk_values") or ""
            care_labels = [
                part.strip() for part in str(crl_raw).split(";") if str(part).strip()
            ]
            care_values = [
                part.strip() for part in str(crv_raw).split(";") if str(part).strip()
            ]
            care_freq = {}
            for label, value in zip(care_labels, care_values):
                try:
                    num = float(str(value).replace(",", ".") or "0")
                except Exception:
                    num = 0.0
                if 0 < num <= 1:
                    num *= 100.0
                care_freq[label] = num

            conn.execute(
                """
                INSERT INTO age_stats
                    (age_group, description, care_recommendations,
                     common_problems, complications_frequency, diseases_by_care)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(age_group) DO UPDATE SET
                    description = excluded.description,
                    care_recommendations = excluded.care_recommendations,
                    common_problems = excluded.common_problems,
                    complications_frequency = excluded.complications_frequency,
                    diseases_by_care = excluded.diseases_by_care,
                    updated_at = CURRENT_TIMESTAMP
                """,
                (
                    age_group,
                    description,
                    care_recommendations,
                    common_problems,
                    json.dumps(comp_freq, ensure_ascii=False),
                    json.dumps(care_freq, ensure_ascii=False),
                ),
            )

    # --- Лист по поведению ---
    seen_behaviors = set()
    if "Behaviors" in wb.sheetnames:
        ws = wb["Behaviors"]
        rows = sheet_to_rows(ws)
        print("[DB] Behaviors rows:", len(rows))
        for row in rows:
            behavior_type = (row.get("behavior_type") or "").strip()
            if not behavior_type:
                continue
            seen_behaviors.add(behavior_type)
            description = (row.get("description") or "") or None
            causes = (row.get("causes") or "") or None
            solutions = (row.get("solutions") or "") or None
            try:
                freq = float(str(row.get("frequency") or "0").replace(",", "."))
            except Exception:
                freq = 0.0
            if 0 < freq <= 1:
                freq *= 100.0

            conn.execute(
                """
                INSERT INTO behavior_stats
                    (behavior_type, description, causes, solutions, frequency, total_cases)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(behavior_type) DO UPDATE SET
                    description = excluded.description,
                    causes = excluded.causes,
                    solutions = excluded.solutions,
                    frequency = excluded.frequency,
                    updated_at = CURRENT_TIMESTAMP
                """,
                (
                    behavior_type,
                    description,
                    causes,
                    solutions,
                    freq,
                    100,  # базовый знаменатель для процента
                ),
            )

    # Аналогично для behavior_stats: удаляем записи, отсутствующие в Excel
    if seen_behaviors:
        placeholders = ",".join("?" for _ in seen_behaviors)
        conn.execute(
            f"DELETE FROM behavior_stats WHERE behavior_type NOT IN ({placeholders})",
            tuple(seen_behaviors),
        )


def rebuild_stats():
    """
    Полный пересчёт статистики breed_stats, age_stats и behavior_stats
    на основе накопленных записей в anamnesis_cases.
    """
    with get_connection() as conn:
        # Считываем все сырые кейсы
        rows = conn.execute(
            "SELECT breed, age_group, symptoms, behaviors FROM anamnesis_cases"
        ).fetchall()

        # Если ещё нет ни одного кейса — сохраняем исходные справочные частоты
        if not rows:
            return

        breed_agg = {}      # {breed_name: {"cases": n, "diseases": {name: count}}}
        age_agg = {}        # {age_group: {"cases": n, "complications": {name: count}}}
        behavior_counts = {}  # {behavior_type: count_cases_with_this_behavior}
        total_cases_overall = len(rows)

        for r in rows:
            breed = (r["breed"] or "").strip()
            age_group = (r["age_group"] or "").strip()
            try:
                symptoms = json.loads(r["symptoms"] or "[]")
            except Exception:
                symptoms = []
            try:
                behaviors = json.loads(r["behaviors"] or "[]")
            except Exception:
                behaviors = []

            # По породам
            if breed:
                b_entry = breed_agg.setdefault(breed, {"cases": 0, "diseases": {}})
                b_entry["cases"] += 1
                for s in symptoms:
                    if not s:
                        continue
                    b_entry["diseases"][s] = b_entry["diseases"].get(s, 0) + 1

            # По возрасту
            if age_group:
                a_entry = age_agg.setdefault(age_group, {"cases": 0, "complications": {}})
                a_entry["cases"] += 1
                for s in symptoms:
                    if not s:
                        continue
                    a_entry["complications"][s] = a_entry["complications"].get(s, 0) + 1

            # По поведению
            # учитываем уникальные типы проблем в рамках одного кейса
            seen_behaviors = set()
            for b in behaviors:
                name = (b or "").strip()
                if not name or name in seen_behaviors:
                    continue
                seen_behaviors.add(name)
                behavior_counts[name] = behavior_counts.get(name, 0) + 1

        # Применяем агрегаты к breed_stats (только для пород, по которым есть данные)
        for breed, data in breed_agg.items():
            cases = max(data["cases"], 1)
            diseases = data["diseases"]
            if not diseases:
                continue
            percents = {k: int(round(v * 100.0 / cases)) for k, v in diseases.items()}
            conn.execute(
                """UPDATE breed_stats
                   SET disease_frequency = ?, total_cases = ?, updated_at = CURRENT_TIMESTAMP
                   WHERE breed_name = ?""",
                (json.dumps(percents, ensure_ascii=False), cases, breed),
            )

        # Применяем агрегаты к age_stats (только для возрастных групп с данными)
        for age_group, data in age_agg.items():
            cases = max(data["cases"], 1)
            comps = data["complications"]
            if not comps:
                continue
            percents = {k: int(round(v * 100.0 / cases)) for k, v in comps.items()}
            conn.execute(
                """UPDATE age_stats
                   SET complications_frequency = ?, total_cases = ?, updated_at = CURRENT_TIMESTAMP
                   WHERE age_group = ?""",
                (json.dumps(percents, ensure_ascii=False), cases, age_group),
            )

        # Применяем агрегаты к behavior_stats (только если есть данные по поведению)
        if total_cases_overall > 0 and behavior_counts:
            for behavior_type, cnt in behavior_counts.items():
                freq = int(round(cnt * 100.0 / total_cases_overall))
                conn.execute(
                    """INSERT INTO behavior_stats (behavior_type, frequency, total_cases)
                       VALUES (?, ?, ?)
                       ON CONFLICT(behavior_type) DO UPDATE SET
                           frequency = excluded.frequency,
                           total_cases = excluded.total_cases,
                           updated_at = CURRENT_TIMESTAMP
                    """,
                    (behavior_type, freq, cnt),
                )
